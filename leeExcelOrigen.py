from openpyxl import load_workbook

#Asignamos la ruta donde estará el archivo
path = '/Users/johamhernandez/Documents/Proyectos Python/ECR_Group/Proyecto/Planillas/p1_sample.xlsx'

#Leemos el archivo
wb = load_workbook(filename = path)

#Leemos todas las columnas de la hoja indicada
#TODO: Asignar el nombre de la hoja como parámetro
sheet_rows = wb['DEPOR'].rows
sheet_columns = wb['DEPOR'].columns
sheet_rows_qty = wb['DEPOR'].max_row
sheet_columns_qty = wb['DEPOR'].max_column
encabezado = []
valores = []
ch = 'A'

print(sheet_rows)
#print(sheet_columns)


ws = wb.active

#Recorremos la hoja de acuerdo a los valores de cada fila y cada columna

for j in range(sheet_rows_qty):
	#i = 0
	sheet_columns = wb['Hoja1'].columns
	for i in sheet_columns:
		#print(j)
		#print(i[j].value)
		if j == 0:
			encabezado.append(i[j].value)

		if j > 0:

			x = chr(ord(ch) + sheet_columns_qty-1)
			#print(x)
			#print(j)
			#Validamos que la ultima columna, tenga valor suspendido
			#para no considerarlo
			valor = ws.cell(row=j, column=sheet_columns_qty).value
			#print(str(valor).upper())
			if not(str(valor).upper()) == 'SUSPENDIDO':
				#print(str(valor))
				#print(ws.cell(row=j, column=sheet_columns_qty).value)
				print(i[j].value)
				valores.append(i[j].value)

#print('encabezado es: ')
#for i in encabezado:
#	print(i)

#print('valores son: ')
#for i in valores:
#	print(i)



#TODO: Guardar los resultados para pasarlos a la creación de un excel
#Es importante evaluar el campo Estatus ya que ese es el que indicará
#Si de debe cargar o no 

# for i in range(1, sheet_ranges + 1):
#     cell_obj = sheet_ranges.cell(row = 1, column = i)
#     print(cell_obj.value)


# for i in sheet_ranges:
# 	ch = 'A'
# 	x = chr(ord(ch) + i)
# 	print(x)



#for k in range(sheet_columns_qty):
 #	ch = 'A'
 #	x = chr(ord(ch) + k)
 #	print(x)


