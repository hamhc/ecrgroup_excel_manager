from openpyxl import load_workbook
from generaExcel import genera_archivo

# TODO: Ruta que venga desde un archivo
path_origen = '/Users/johamhernandez/Documents/Proyectos Python/ECR_Group/Proyecto/Planillas/p1_sample.xlsx'
path_destino = '/Users/johamhernandez/Documents/Proyectos Python/ECR_Group/Proyecto/Results/'

def lee_excel(nombre_cliente, estado_empleado):
	# Leemos el archivo
	wb = load_workbook(filename = path_origen)

	#asignamos contenido de hoja por nombre
	hoja = wb[nombre_cliente]

	#Buscamos valor del campo estatus
	#estatus = hoja['Z3']
	#print(estatus.value)

	#Obtenemos las celdas que tienen data en el excel
	#rango_celdas = hoja['A1':'Z4']

	#Obtenemos total de columnas
	sheet_columns_qty = hoja.max_column
	#Obtenemos total de filas
	sheet_rows_qty = hoja.max_row

	#Asignamos al rango desde A1 hasta Z + el total de filas que hay en la hoja
	rango_celdas = hoja['A1':'Z'+str(sheet_rows_qty)]

	encabezado = []
	#contenido_tmp = []
	contenido = []

	#Counter que se utilizada cuando se guarda el contenido para no insertar lineas en blanco
	counter = 0

	#print(sheet_rows_qty)
	#print(sheet_columns_qty)
	#print(rango_celdas)
	#print(rango_celdas[0][0].value)

	for j in range(0,sheet_rows_qty):

		contenido_tmp = []
		for i in range(0,sheet_columns_qty):
			#primer valor = Filas
			#segundo valor = columnas
			#print(rango_celdas[0][i].value)

			#print(rango_celdas[j][sheet_columns_qty-1].value)
			valor = str(rango_celdas[j][sheet_columns_qty-1].value)
			#print(valor)

			if(j == 0):
				#print(rango_celdas[0][i].value)
				encabezado.append(rango_celdas[j][i].value)

			elif(j > 0):
				#if not ((valor.upper()) == 'SUSPENDIDO'):
				if ((valor.upper()) == estado_empleado.upper()):
					#print(rango_celdas[j][i].value)
					#Agregamos uno al counter para agregar las filas sin espacios en blanco
					counter += 1
					#contenido_tmp.append(rango_celdas[counter][i].value)
					contenido_tmp.append(rango_celdas[j][i].value)

		if(j > 0):
			if ((valor.upper()) == estado_empleado.upper()):
				contenido.append(contenido_tmp)

	print(encabezado)
	print(contenido)

	genera_archivo(nombre_cliente, nombre_cliente+'_'+'18042021', path_destino, encabezado, contenido)