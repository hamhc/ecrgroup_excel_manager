#from openpyxl import load_workbook
import openpyxl, os, datetime

class ExcelManager:
    def __init__(self, cliente, estado_empleado):
        self.cliente = cliente
        self.estado_empleado = estado_empleado
        self.path_origen = '/Users/johamhernandez/Documents/Proyectos Python/ECR_Group/Proyecto/Planillas/p1_sample.xlsx'
        self.path_destino = '/Users/johamhernandez/Documents/Proyectos Python/ECR_Group/Proyecto/Results/'
        self.define_fecha()
        self.nombre_archivo = cliente + '_' + self.date + '.xlsx'

    #def leeExcel(self):
    def lee_excel(self):
        """ Leemos el archivo origen de excel, que tiene matriz con informacion completa de todos los clientes
            separamos el contenido leido en 2 listas: encabezado y contenido
        """
        wb = openpyxl.load_workbook(filename=self.path_origen)

        # asignamos contenido de hoja por nombre
        hoja = wb[self.cliente]

        # Buscamos valor del campo estatus
        # estatus = hoja['Z3']
        # print(estatus.value)

        # Obtenemos las celdas que tienen data en el excel
        # rango_celdas = hoja['A1':'Z4']

        # Obtenemos total de columnas
        sheet_columns_qty = hoja.max_column
        # Obtenemos total de filas
        sheet_rows_qty = hoja.max_row

        # Asignamos al rango desde A1 hasta Z + el total de filas que hay en la hoja
        rango_celdas = hoja['A1':'Z' + str(sheet_rows_qty)]

        encabezado = []
        # contenido_tmp = []
        contenido = []

        # Counter que se utilizada cuando se guarda el contenido para no insertar lineas en blanco
        counter = 0

        # print(sheet_rows_qty)
        # print(sheet_columns_qty)
        # print(rango_celdas)
        # print(rango_celdas[0][0].value)

        for j in range(0, sheet_rows_qty):

            contenido_tmp = []
            for i in range(0, sheet_columns_qty):
                # primer valor = Filas
                # segundo valor = columnas
                # print(rango_celdas[0][i].value)

                # print(rango_celdas[j][sheet_columns_qty-1].value)
                valor = str(rango_celdas[j][sheet_columns_qty - 1].value)
                # print(valor)

                if (j == 0):
                    # print(rango_celdas[0][i].value)
                    encabezado.append(rango_celdas[j][i].value)

                elif (j > 0):
                    # if not ((valor.upper()) == 'SUSPENDIDO'):
                    if ((valor.upper()) == self.estado_empleado.upper()):
                        # print(rango_celdas[j][i].value)
                        # Agregamos uno al counter para agregar las filas sin espacios en blanco
                        counter += 1
                        # contenido_tmp.append(rango_celdas[counter][i].value)
                        contenido_tmp.append(rango_celdas[j][i].value)

            if (j > 0):
                if ((valor.upper()) == self.estado_empleado.upper()):
                    contenido.append(contenido_tmp)
        print(encabezado)
        print(contenido)
        self.encabezado = encabezado
        self.contenido = contenido

        #self.genera_archivo(encabezado, contenido)

    def genera_archivo(self):
        """Generamos archivo excel de forma logica, la hoja tendrá el nombre que corresponde al cliente indicado,
        también validamos el estado deseado"""
        wb = openpyxl.Workbook()

        #Obtenemos cual es la hoja Activa del archivo
        active = wb.active

        #Asignamos el nombre a la hoja generada
        active.title = self.cliente

        #Asignamos la hoja Deport (del libro) a una variable
        sheet = wb[self.cliente]
        #Asignamos valores a las celdas A1 y A2
        #sheet['A1'] = 42
        #sheet['A2'] = 'Caraotas'
        #for i in encabezado:
        sheet.append(self.encabezado)

        for i in self.contenido:
            sheet.append(i)

        #guarda_archivo(nombre_archivo, '/Users/johamhernandez/Documents/Proyectos Python/ECR_Group/Proyecto/Results', wb)
        #guarda_archivo(nombre_archivo, path, wb)
        self.wb_object = wb

    def guarda_archivo(self):
        """Funcion para guardar archivo"""

        # Nos cambiamos de directorio para guardar el archivo
        os.chdir(self.path_destino)

        # Guardamos el achivo excel después de cambiarnos de ruta
        # NOTA: si no lo guardamos, solo existe en memoria
       self.wb_object.save(self.nombre_archivo)

    def define_fecha(self):
        """Definimos un string de fecha para concantenarle al nombre del archivo
        el formato es: DDMMYYYY_HHMMSS"""
        date = datetime.datetime.now()
        self.date = date.strftime("%d%m%Y_%H%M%S")

    def ejecutor(self):
        """Ejecutor que permite leer archivo origen, generar nuevo archivo logico y guardar archivo excel fisico"""
        self.lee_excel()
        self.genera_archivo()
        self.guarda_archivo()

p1 = ExcelManager('DEPOR', 'suspendido')
p1.ejecutor()